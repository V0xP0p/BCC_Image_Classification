# utils.py contains function and classes used in preprocessing, modelling and visualization of
# machine learning models.

import numpy as np
import pandas as pd
import os
# from os.path import exists
import pathlib
import inspect
import tensorflow as tf
import json
# import pyyaml
from openpyxl import load_workbook, Workbook
from sklearn.model_selection import StratifiedKFold, train_test_split, StratifiedGroupKFold
from sklearn.metrics import confusion_matrix, ConfusionMatrixDisplay, roc_curve, auc
from sklearn.metrics import precision_score, recall_score, f1_score, balanced_accuracy_score
from sklearn.metrics import roc_auc_score, mean_squared_error, accuracy_score


class ImageDataset:

    AUTOTUNE = tf.data.AUTOTUNE
    label_counter = 0

    def __init__(self, filepath, entries=None, labels=None, image_type='jpg', search_type='*'):
        self.filepath = pathlib.Path(filepath)
        self.str_filepath = filepath
        self.entries = entries

        if labels is None:
            self.no_labels = True
        else:
            self.no_labels = False

        if not self.filepath.exists():
            raise ValueError(f"{self.filepath} is not a valid path.")
        elif not self.filepath.is_dir():
            raise ValueError(f"{self.filepath} is not a directory.")

        if self.no_labels and search_type == '*':
            self.search_type = '*/*'
            self.path_list = tf.data.Dataset.list_files(str(self.filepath / self.search_type), shuffle=False)
        else:
            self.search_type = search_type
            # entries = os.listdir(self.str_filepath)
            # entries.sort(key=len)
            img_paths = []
            for entry in self.entries:
                img_paths.append(os.path.join(filepath, entry))
            self.path_list = tf.data.Dataset.from_tensor_slices(img_paths)

        if self.no_labels:
            self.labels = np.array(sorted([item.name for item in self.filepath.glob('*') if item.name != "LICENSE.txt"]))
        else:
            if type(labels) == pd.DataFrame:
                self.labels = labels.to_numpy()
            else:
                self.labels = labels

        self.image_type = image_type

        self.image_count = len(list(self.filepath.glob(self.search_type)))

        self.dataset = tf.data.Dataset



    # Instance method that provides a short
    # description of the class instance created.
    def __str__(self):
        return f"This is an instance of the image dataset in {self.filepath}"

    def get_path_list(self):
        return self.path_list

    def get_dataset(self):
        return self.dataset

    def set_dataset(self, ds):
        self.dataset = ds

    def _get_labels(self, filepath):

        if self.no_labels:
            # Convert the path to a list of path components
            parts = tf.strings.split(filepath, os.path.sep)
            # The second to last is the class-directory
            one_hot = parts[-2] == self.labels
            # Integer encode the label
            return tf.argmax(one_hot)
        else:
            label = self.labels[self.label_counter]
            self.label_counter += 1
            return label

    def _decode_image(self, img, resize):
        if self.image_type == 'jpg':
            # Convert the compressed string to a 3D uint8 tensor
            img = tf.io.decode_jpeg(img, channels=3)

        else:
            raise TypeError(f"No compatibility for type: {self.image_type} implemented yet.")

        if resize is not None:
            # Resize the image to the desired size
            return tf.image.resize(img, resize)
        else:
            return img

    def create_image_label_set(self, filepath):

        label = self._get_labels(filepath)
        # Load the raw data from the file as a string
        img = tf.io.read_file(filepath)
        img = self._decode_image(img, resize=[224, 224])

        return img, label

    def create_image_dataset(self, num_parallel_calls=AUTOTUNE):

        self.set_dataset(self.path_list.map(self.create_image_label_set, num_parallel_calls))

        return self.get_dataset()

    @staticmethod
    def performance_config(ds,
                           cache=True,
                           shuffle=True,
                           shuffle_buff_size=300,
                           batch_size=32,
                           augment=False,
                           augmentation=None,
                           prefetch=True,
                           prefetch_buff_size=AUTOTUNE):

        if cache:
            ds = ds.cache()
        if shuffle:
            ds = ds.shuffle(buffer_size=shuffle_buff_size)
        if batch_size is not None:
            ds = ds.batch(batch_size)
        if augment:
            ds = ds.map(lambda x, y: (augmentation(x, training=True), y),
                        num_parallel_calls=prefetch_buff_size)
        if prefetch:
            ds = ds.prefetch(buffer_size=prefetch_buff_size)

        return ds


def _check_args(params_to_check):
    def inner(func):
        def wrapper(*args, **kwargs):
            for param_to_check in params_to_check:
                if param_to_check not in inspect.signature(func).parameters:
                    raise TypeError(f"Parameter '{param_to_check}' must be passed")
                return func(*args, **kwargs)
        return wrapper
    return inner


class Logger:

    def __init__(self, verbose=True):
        self.verbose = verbose

    def set_verbosity(self, verbose=True):
        self.verbose = verbose

    @_check_args(params_to_check=['filename', 'values'])
    def to_excel(self, filename, values):

        if os.path.exists(filename):

            wb = load_workbook(filename)
            ws = wb.active
            title_row = ws[ws.min_row]
            labels = {key.value: value for (key, value) in zip(title_row, range(1, len(title_row)+1))}
            # labels = [cell.value for cell in ws[ws.min_row]]
            # # if len(labels) != len(values) + 1:
            # #     raise ValueError(f"Expected {len(labels) - 1} arguments and instead {len(values)} were passed.")
            new_row_index = ws.max_row + 1

            experiment_num = new_row_index - 1

            _ = ws.cell(new_row_index, 1, f"Experiment {experiment_num}")

            for label, value in values.items():
                if label in labels:
                    _ = ws.cell(new_row_index, labels[label], value)
                else:
                    col_idx = ws.max_column + 1
                    _ = ws.cell(ws.min_row, col_idx, label)
                    _ = ws.cell(new_row_index, col_idx, value)

            wb.save(filename)

        else:
            wb = Workbook()
            ws = wb.active
            labels = ["Experiment No"] + list(values.keys())
            ws.append(labels)
            row = [f"Experiment 1"]
            row.extend(values.values())
            ws.append(row)
            wb.save(filename)

            experiment_num = 1

        if self.verbose is True:
            print(f"\nLogged the following variables:")
            print(30 * '-')
            for label, value in values.items():
                print(f"{label}: {value:.2f}")

        return experiment_num

    @_check_args(params_to_check=['filename', 'values'])
    def to_json(self, filename, values):
        json.dump(values, open(filename, 'w'))


def get_stemmed_filename(filename, cast_type=None):
    """Function that accepts a filename and extracts the filename
    the final path component without its suffix.

    :param filename: the filename we want to stem
    :param cast_type: the type we want to cast the result to.
                    i.e. int, str, float etc.

    :return: stemmed filename in the type specified
    """

    if cast_type is None:
        return pathlib.Path(filename).stem
    else:
        return cast_type(pathlib.Path(filename).stem)


def initialize_function(func_name, *args, **kwargs):
    """
    Call a function with specific parameters initialized otherwise
    it uses the defaults for that function.

    :param func_name: <string> the function we want to call.
    :param kwargs: named arguments consistent with the function called and initialized.
    :return: the function initialized with the new arguments.
    """
    method_params = {}

    func = eval(func_name)

    for parameter in inspect.signature(func).parameters:

        if parameter in kwargs:
            param_value = kwargs.get(parameter)
            method_params[parameter] = param_value

    return func(*args, **method_params)


def split_data(data, labels, split_strategy='StratifiedKFold', **kwargs):
    """
    Splits the data to train-test sets depending on the specified strategy

    :param data: original unsplit data
    :param labels: original unsplit labels
    :param split_strategy: strategy to be followed for the split,
            currently supports: ['StratifiedKFold', 'train_test_split']
    :param kwargs: named arguments

    :return: depending on the method either an array containing the different splits or
            the X_train, X_test, Y_train, Y_test arrays.
    """

    if split_strategy == 'StratifiedKFold':
        split_method = initialize_function(split_strategy, **kwargs)
        return zip(split_method.split(data, labels), range(split_method.n_splits))
    elif split_strategy == 'train_test_split':
        return initialize_function(split_strategy, data, labels, **kwargs)
    elif split_strategy == 'StratifiedGroupKFold':
        split_method = initialize_function(split_strategy, **kwargs)
        return zip(split_method.split(data, labels), range(split_method.n_splits))


def get_classifier():
    pass


def compute_roc_auc(index, X, y, clf):
    y_predict = clf.predict_proba(X.iloc[index])[:, 1]
    fpr, tpr, thresholds = roc_curve(y.iloc[index], y_predict)
    auc_score = auc(fpr, tpr)
    return fpr, tpr, auc_score


def get_metrics(y_true, y_pred, index=None, data=None, labels=None, classifier=None, verbose=True, filename=None):

    roc_flag = (index is not None) and (data is not None) and (classifier is not None) and (labels is not None)

    precision = precision_score(y_true, y_pred, average='weighted', zero_division=0)
    recall = recall_score(y_true, y_pred, average='weighted', zero_division=0)
    f1 = f1_score(y_true, y_pred, average='weighted')
    bacc = balanced_accuracy_score(y_true, y_pred)
    if roc_flag:
        _, _, roc = compute_roc_auc(index, data, labels, classifier)
    else:
        pass
        # raise ValueError(f'For ROC score to be evaluated index,
        # data and classifier fields must be explicitly passed.\n'
        #                  f'Current values:\n'
        #                  f'index= {index}\n'
        #                  f'data= {data}\n'
        #                  f'classifier= {classifier}')
    acc = accuracy_score(y_true, y_pred)

    if verbose:
        print(f"Precision= {precision}")
        print(f"Recall= {recall}")
        print(f"F1 score= {f1}")
        print(f"Balanced Accuracy {bacc}")
        if roc_flag:
            print(f"ROC score {roc}")
        print(f"Accuracy {acc}")

    if filename is not None:
        print(f"Filename is {filename}")
        if os.path.exists(filename):
            wb = load_workbook(filename)
            ws = wb.active
            new_row_index = ws.max_row
            row = [f"Experiment {new_row_index}", precision, recall, f1, acc, bacc, roc]
            ws.append(row)
            wb.save(filename)

        else:
            wb = Workbook()
            ws = wb.active

            rows = [[f"Experiment Number", "Precision", "Recall", "F1 score", "Accuracy",
                     "Balanced Accuracy", "ROC-AUC Score"],
                    [f"Experiment 1", precision, recall, f1, acc, bacc, roc]]
            for row in rows:
                ws.append(row)
            wb.save(filename)

    return precision, recall, f1, acc, bacc


def save_metrics(filename, metrics=None, labels=None):

    if metrics is not None:
        if os.path.exists(filename):
            wb = load_workbook(filename)
            ws = wb.active
            previous_row = [metric.value for metric in ws[ws.min_row]]
            if len(previous_row) != len(metrics)+1:
                raise ValueError(f"Expected {len(previous_row)} arguments and instead {len(metrics)} were passed.")
            new_row_index = ws.max_row + 1
            row = [f"Experiment {new_row_index-1}"]
            # metrics = [metric[0] for metric in metrics]
            row.extend(metrics)
            ws.append(row)
            wb.save(filename)
        else:
            wb = Workbook()
            ws = wb.active
            if labels is None:
                raise ValueError(f"For first time initialization variables 'metrics' and 'labels' must be "
                                 f"passed explicitly."
                                 f"Current Values:"
                                 f"labels= {labels}")

            metrics = [f"Experiment 1"] + metrics   #[metric[0] for metric in metrics]

            labels = ["Experiment Number"] + labels
            rows = [labels, metrics]
            for row in rows:
                ws.append(row)
            wb.save(filename)
    else:
        raise ValueError(f"For first time initialization variables 'metrics' and 'labels' must be "
                         f"passed explicitly."
                         f"Current Values:"
                         f"metrics= {metrics}")


def read_metrics(filename, exp_num=None):

    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active

        max_idx = ws.max_row
        min_idx = ws.min_row

        labels = [label.value for label in ws[1]]
        # labels = labels[1:]

        if exp_num is None:
            # if exp_num is none retrieve the metrics for the last experiment
            metrics = [metric.value for metric in ws[max_idx]]
            metrics_dict = dict(zip(labels, metrics))
        else:
            exp_num_type = type(exp_num)
            if exp_num_type == list:
                metrics_dict = dict()
                for experiment in exp_num:
                    metrics = [metric.value for metric in ws[experiment + 1]]
                    metrics_dict[f"Experiment {experiment}"] = dict(zip(labels[1:], metrics[1:]))
            elif exp_num_type == int:
                metrics = [metric.value for metric in ws[exp_num + 1]]
                metrics_dict = dict(zip(labels, metrics))
            else:
                raise TypeError(f"The value of the argument 'exp_num' must be of type 'list' or 'int'"
                                f"Currently exp_num is of type: {exp_num_type}")

        return max_idx-1, metrics_dict

    else:
        return 1, None


def balance_data_by_reduction(data):
    """
    Balances a dataset by reducing its rows with respect to the class
    that has the least labels. Not suited for highly unbalanced datasets.

    :param data: a pandas dataframe of the labels in one-hot format.
    :return: the sorted indices of the reduced dataset
    """

    data_len = len(data)
    num_classes = data.shape[1]

    class_freq = [ x/data_len for x in list(np.sum(data, axis=0))]

    least_class = np.argmin(class_freq)
    class_size = np.sum(data.iloc[:, least_class])

    indices = []
    indices.extend(data[data.iloc[:, least_class] == 1].index)

    for pred_class in range(num_classes):
        if pred_class != least_class:
            indices.extend(np.random.choice(data[data.iloc[:, pred_class] == 1].index, class_size, replace=False))

    indices.sort()

    return indices


def reorder_image_labels(filepath,
                         labels,
                         index_start=1,
                         valid_extensions=('.jpg', '.png', '.bmp', '.gif')):
    """
    Reorders the labels of an image dataset sorted with the
    alphanumeric order of image files found in the directory.
    Assumes images named '1.extension', '2.extension' etc. where
    the number corresponds with the index of the correct label
    depending on index_start.

    :param filepath: the filepath where the images reside.
    :param labels: a list of integers containing the labels
                    for each image in ascending order.
    :param index_start: integer specifies the starting index of
                        the images. Defaults to 1.
    :param valid_extensions: list of strings, contains the valid file extensions
                            i.e. ['.png', '.jpg']
    :return: a list of labels ordered based on the alphanumeric
            file order.
    """

    image_paths = []

    # iterate through all subfolders and files in
    # the filepath.
    for _, _, files in os.walk(filepath):
        # iterate through all files in the directory.
        for file in files:
            # check if the file extension is in valid extensions.
            if os.path.splitext(file)[1].lower() in valid_extensions:
                # add only valid filetypes.
                image_paths.append(file)

    reordered_labels = []
    # iterate through all the files found.
    for image in image_paths:
        # update the index of the label based on the filename
        # taking into account the index start.
        position = int(os.path.splitext(image)[0]) - index_start
        # add the reordered label to a list.
        reordered_labels.append(labels[position])

    return reordered_labels


def one_hot_to_integers(df):

    return pd.DataFrame(df.apply(np.argmax, axis=1))


if __name__ == '__main__':
    print("This is the main module")
else:
    print('Main module did not execute')

